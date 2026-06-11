import os
import base64
import json
import re
import io
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import FastAPI, File, UploadFile, Depends, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from pydantic import BaseModel
from dotenv import load_dotenv
import httpx

from database import get_db, init_db, Arbitro, Partido, AsignacionPartido, Reemplazo, Jornada  # noqa: E402

load_dotenv()
init_db()

app = FastAPI(title="Gestor de Árbitros")

@app.middleware("http")
async def no_cache_static(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    return response

app.mount("/static", StaticFiles(directory="static"), name="static")

MATCH_DURATION_HOURS = 2  # Duración estimada de un partido


# ─── Schemas ────────────────────────────────────────────────────────────────

class ArbitroCreate(BaseModel):
    nombre: str
    categoria: str
    telefono: Optional[str] = None


class ArbitroUpdate(BaseModel):
    telefono: Optional[str] = None
    nombre: Optional[str] = None


class PartidoManual(BaseModel):
    numero: str
    equipo_local: str
    equipo_visitante: str
    competicion: str
    fecha_jornada: str
    estadio: str
    fecha_hora: str  # "DD.MM.YYYY HH:MM"
    numero_partido: str
    departamento: str
    ciudad: str
    imagen_url: Optional[str] = None
    asignaciones: List[dict]  # [{"rol": "Árbitro", "nombre": "..."}]


# ─── Helpers ────────────────────────────────────────────────────────────────

def parse_fecha(fecha_str: str) -> Optional[datetime]:
    """Parse DD.MM.YYYY HH:MM or DD.MM.YYYY"""
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(fecha_str.strip(), fmt)
        except ValueError:
            continue
    return None


def detectar_conflictos(partido_id: int, db: Session):
    """Detecta conflictos de horario para todos los árbitros de un partido."""
    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido or not partido.fecha_hora:
        return []

    conflictos = []
    fin_partido = partido.fecha_hora + timedelta(hours=MATCH_DURATION_HOURS)

    for asig in partido.asignaciones:
        # Saltar si el árbitro fue eliminado (FK huérfana)
        if asig.arbitro is None:
            continue

        # Buscar otros partidos del mismo árbitro
        otras = (
            db.query(AsignacionPartido)
            .filter(
                AsignacionPartido.arbitro_id == asig.arbitro_id,
                AsignacionPartido.partido_id != partido_id,
            )
            .all()
        )

        for otra in otras:
            otro_partido = otra.partido
            if not otro_partido.fecha_hora:
                continue

            otro_fin = otro_partido.fecha_hora + timedelta(hours=MATCH_DURATION_HOURS)

            # Conflicto si horarios se solapan O mismo día
            mismo_dia = partido.fecha_hora.date() == otro_partido.fecha_hora.date()
            solapan = (
                partido.fecha_hora < otro_fin and fin_partido > otro_partido.fecha_hora
            )

            if mismo_dia or solapan:
                conflictos.append(
                    {
                        "arbitro_id": asig.arbitro_id,
                        "arbitro_nombre": asig.arbitro.nombre,
                        "rol_en_este": asig.rol,
                        "partido_conflicto_id": otro_partido.id,
                        "partido_conflicto_num": otro_partido.numero,
                        "equipos_conflicto": f"{otro_partido.equipo_local} vs {otro_partido.equipo_visitante}",
                        "fecha_conflicto": otro_partido.fecha_hora.strftime("%d.%m.%Y %H:%M") if otro_partido.fecha_hora else "",
                        "estadio_conflicto": otro_partido.estadio or "",
                        "ciudad_conflicto": otro_partido.ciudad or "",
                        "competicion_conflicto": otro_partido.competicion or "",
                        "rol_en_conflicto": otra.rol,
                        "tipo": "solapamiento" if solapan else "mismo_dia",
                    }
                )

    return conflictos


def sugerir_reemplazos(arbitro_id: int, partido_id: int, db: Session):
    """Sugiere árbitros disponibles de la misma categoría."""
    arbitro = db.query(Arbitro).filter(Arbitro.id == arbitro_id).first()
    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not arbitro or not partido or not partido.fecha_hora:
        return []

    fin_partido = partido.fecha_hora + timedelta(hours=MATCH_DURATION_HOURS)

    # Todos los árbitros activos excepto el que tiene conflicto
    candidatos = (
        db.query(Arbitro)
        .filter(Arbitro.activo == True, Arbitro.id != arbitro_id)
        .all()
    )

    disponibles = []
    for candidato in candidatos:
        ocupado = False
        for asig in candidato.asignaciones:
            p = asig.partido
            if not p.fecha_hora:
                continue
            p_fin = p.fecha_hora + timedelta(hours=MATCH_DURATION_HOURS)
            # Solo excluir si los horarios SE SOLAPAN realmente
            solapan = partido.fecha_hora < p_fin and fin_partido > p.fecha_hora
            if solapan:
                ocupado = True
                break
        if not ocupado:
            # Partidos ese mismo día (sin solapamiento)
            partidos_ese_dia = []
            for asig in candidato.asignaciones:
                p2 = asig.partido
                if p2.fecha_hora and p2.fecha_hora.date() == partido.fecha_hora.date():
                    partidos_ese_dia.append({
                        "hora": p2.fecha_hora.strftime("%H:%M"),
                        "estadio": p2.estadio or "",
                        "ciudad": p2.ciudad or "",
                        "rol": asig.rol,
                    })
            disponibles.append({
                "id": candidato.id,
                "nombre": candidato.nombre,
                "categoria": candidato.categoria,
                "telefono": candidato.telefono or "",
                "aviso_mismo_dia": len(partidos_ese_dia) > 0,
                "partidos_ese_dia": partidos_ese_dia,
            })

    return disponibles


def extraer_datos_imagen(imagen_b64: str, media_type: str) -> dict:
    """Usa Gemini Vision para extraer datos del partido desde la imagen."""
    prompt = """Analiza esta imagen de un sistema de gestión de partidos de fútbol y extrae los datos indicados.

Responde ÚNICAMENTE con un JSON válido con esta estructura exacta (sin texto adicional):
{
  "equipo_local": "nombre del equipo de la izquierda",
  "equipo_visitante": "nombre del equipo de la derecha",
  "competicion": "texto del campo Competición",
  "estadio": "texto del campo Estadio",
  "fecha_hora": "fecha y hora en formato DD.MM.YYYY HH:MM",
  "ciudad": "texto del campo Ciudad / Municipio",
  "asignaciones": [
    {"rol": "Árbitro", "nombre": "APELLIDO NOMBRE (CATEGORIA)"},
    {"rol": "1° árbitro asistente", "nombre": "APELLIDO NOMBRE (CATEGORIA)"},
    {"rol": "2° árbitro asistente", "nombre": "APELLIDO NOMBRE (CATEGORIA)"},
    {"rol": "Cuarto árbitro", "nombre": "APELLIDO NOMBRE (CATEGORIA)"}
  ]
}

IMPORTANTE:
- En asignaciones incluye SOLO los roles que tengan nombre asignado (no vacíos)
- Copia los nombres EXACTAMENTE como aparecen en la imagen incluyendo la categoría entre paréntesis
- Si un campo no aparece, usa cadena vacía ""
- El JSON debe ser válido y completo"""

    api_key = os.getenv("OPENROUTER_API_KEY", "")
    with httpx.Client(timeout=60.0) as client:
        resp = client.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:8000",
                "X-Title": "Gestor de Arbitros",
            },
            json={
                "model": "nvidia/nemotron-nano-12b-v2-vl:free",
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{imagen_b64}"}},
                        {"type": "text", "text": prompt},
                    ],
                }],
            },
        )
    data = resp.json()
    if "error" in data:
        raise Exception(str(data["error"]))
    text = data["choices"][0]["message"]["content"].strip()
    # Extraer JSON si viene con markdown
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        text = match.group(0)
    return json.loads(text)


# ─── Rutas principales ───────────────────────────────────────────────────────

@app.get("/")
async def index():
    return FileResponse("templates/index.html")


# ─── API Jornadas ─────────────────────────────────────────────────────────────

class JornadaCreate(BaseModel):
    nombre: str


@app.get("/api/jornadas")
def listar_jornadas(db: Session = Depends(get_db)):
    return [{"id": j.id, "nombre": j.nombre, "total_partidos": len(j.partidos)}
            for j in db.query(Jornada).order_by(Jornada.creado_en.desc()).all()]


@app.post("/api/jornadas")
def crear_jornada(data: JornadaCreate, db: Session = Depends(get_db)):
    j = Jornada(nombre=data.nombre.strip())
    db.add(j)
    db.commit()
    db.refresh(j)
    return {"id": j.id, "nombre": j.nombre}


@app.delete("/api/jornadas/{jornada_id}")
def eliminar_jornada(jornada_id: int, db: Session = Depends(get_db)):
    j = db.query(Jornada).filter(Jornada.id == jornada_id).first()
    if not j:
        raise HTTPException(status_code=404, detail="Jornada no encontrada")
    db.delete(j)
    db.commit()
    return {"ok": True}


# ─── API Árbitros ────────────────────────────────────────────────────────────

@app.get("/api/debug-key")
def debug_key():
    key = os.getenv("OPENROUTER_API_KEY", "")
    return {"tiene_clave": bool(key), "primeros_chars": key[:12] if key else "VACÍA"}



@app.get("/api/arbitros")
def listar_arbitros(db: Session = Depends(get_db)):
    arbitros = db.query(Arbitro).filter(Arbitro.activo == True).order_by(Arbitro.nombre).all()
    result = []
    for a in arbitros:
        conteo = {}
        total = 0
        for asig in a.asignaciones:
            rol = asig.rol
            conteo[rol] = conteo.get(rol, 0) + 1
            total += 1
        result.append({
            "id": a.id,
            "nombre": a.nombre,
            "telefono": a.telefono or "",
            "total_partidos": total,
            "por_rol": conteo,
        })
    # Ordenar por total de partidos descendente
    result.sort(key=lambda x: x["total_partidos"], reverse=True)
    return result


@app.post("/api/arbitros")
def crear_arbitro(data: ArbitroCreate, db: Session = Depends(get_db)):
    arbitro = Arbitro(nombre=data.nombre.upper().strip(), categoria=data.categoria.upper().strip(), telefono=data.telefono)
    db.add(arbitro)
    db.commit()
    db.refresh(arbitro)
    return arbitro


@app.patch("/api/arbitros/{arbitro_id}")
def actualizar_arbitro(arbitro_id: int, data: ArbitroUpdate, db: Session = Depends(get_db)):
    arbitro = db.query(Arbitro).filter(Arbitro.id == arbitro_id).first()
    if not arbitro:
        raise HTTPException(status_code=404, detail="Árbitro no encontrado")
    if data.telefono is not None:
        arbitro.telefono = data.telefono.strip()
    if data.nombre is not None:
        nuevo = data.nombre.strip().upper()
        if not nuevo:
            raise HTTPException(status_code=400, detail="El nombre no puede estar vacío")
        arbitro.nombre = nuevo
    db.commit()
    return {"id": arbitro.id, "telefono": arbitro.telefono, "nombre": arbitro.nombre}


@app.delete("/api/arbitros/{arbitro_id}")
def eliminar_arbitro(arbitro_id: int, db: Session = Depends(get_db)):
    arbitro = db.query(Arbitro).filter(Arbitro.id == arbitro_id).first()
    if not arbitro:
        raise HTTPException(status_code=404, detail="Árbitro no encontrado")
    arbitro.activo = False
    db.commit()
    return {"ok": True}


# ─── API Partidos ─────────────────────────────────────────────────────────────

@app.get("/api/partidos")
def listar_partidos(jornada_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(Partido)
    if jornada_id:
        q = q.filter(Partido.jornada_id == jornada_id)
    partidos = q.order_by(Partido.fecha_hora).all()
    result = []
    for p in partidos:
        conflictos_raw = detectar_conflictos(p.id, db)
        # Filtrar conflictos que ya tienen reemplazo asignado
        conflictos = []
        for c in conflictos_raw:
            reemplazo = db.query(Reemplazo).filter(
                Reemplazo.arbitro_original_id == c["arbitro_id"],
                Reemplazo.partido_id.in_([p.id, c["partido_conflicto_id"]])
            ).first()
            if not reemplazo:
                conflictos.append(c)
        result.append(
            {
                "id": p.id,
                "numero": p.numero,
                "numero_partido": p.numero_partido or "",
                "jornada_id": p.jornada_id,
                "jornada_nombre": p.jornada.nombre if p.jornada else "",
                "equipo_local": p.equipo_local,
                "equipo_visitante": p.equipo_visitante,
                "competicion": p.competicion,
                "fecha_jornada": p.fecha_jornada or "",
                "estadio": p.estadio,
                "fecha_hora": p.fecha_hora.strftime("%d.%m.%Y %H:%M") if p.fecha_hora else "",
                "departamento": p.departamento or "",
                "ciudad": p.ciudad,
                "imagen_url": p.imagen_url or "",
                "asignaciones": [
                    {
                        "id": a.id,
                        "rol": a.rol,
                        "nombre": a.arbitro.nombre if a.arbitro else "—",
                        "categoria": a.arbitro.categoria if a.arbitro else "",
                        "arbitro_id": a.arbitro_id,
                        "confirmado": a.confirmado or False,
                        "reemplazo": (lambda r: r.arbitro_reemplazo.nombre if r and r.arbitro_reemplazo else None)(
                            db.query(Reemplazo).filter(
                                Reemplazo.arbitro_original_id == a.arbitro_id,
                                Reemplazo.partido_id == p.id
                            ).first()
                        ),
                    }
                    for a in p.asignaciones if a.arbitro_id is not None
                ],
                "conflictos": conflictos,
                "tiene_conflicto": len(conflictos) > 0,
            }
        )
    return result


@app.get("/api/partidos/{partido_id}")
def obtener_partido(partido_id: int, db: Session = Depends(get_db)):
    p = db.query(Partido).filter(Partido.id == partido_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Partido no encontrado")
    conflictos = detectar_conflictos(p.id, db)
    return {
        "id": p.id,
        "numero": p.numero,
        "equipo_local": p.equipo_local,
        "equipo_visitante": p.equipo_visitante,
        "competicion": p.competicion,
        "fecha_jornada": p.fecha_jornada,
        "estadio": p.estadio,
        "fecha_hora": p.fecha_hora.strftime("%d.%m.%Y %H:%M") if p.fecha_hora else "",
        "numero_partido": p.numero_partido,
        "departamento": p.departamento,
        "ciudad": p.ciudad,
        "asignaciones": [
            {"rol": a.rol, "nombre": a.arbitro.nombre if a.arbitro else "—", "categoria": a.arbitro.categoria if a.arbitro else "", "arbitro_id": a.arbitro_id}
            for a in p.asignaciones if a.arbitro_id is not None
        ],
        "conflictos": conflictos,
        "tiene_conflicto": len(conflictos) > 0,
    }


@app.post("/api/partidos")
def guardar_partido(data: PartidoManual, jornada_id: Optional[int] = None, db: Session = Depends(get_db)):
    # Verificar si ya existe
    existente = db.query(Partido).filter(Partido.numero == data.numero).first()
    if existente:
        # Actualizar asignaciones
        db.query(AsignacionPartido).filter(AsignacionPartido.partido_id == existente.id).delete()
        existente.equipo_local = data.equipo_local
        existente.equipo_visitante = data.equipo_visitante
        existente.competicion = data.competicion
        existente.fecha_jornada = data.fecha_jornada
        existente.estadio = data.estadio
        existente.fecha_hora = parse_fecha(data.fecha_hora)
        existente.numero_partido = data.numero_partido
        existente.departamento = data.departamento
        existente.ciudad = data.ciudad
        if data.imagen_url:
            existente.imagen_url = data.imagen_url
        partido = existente
    else:
        partido = Partido(
            numero=data.numero,
            jornada_id=jornada_id,
            equipo_local=data.equipo_local,
            equipo_visitante=data.equipo_visitante,
            competicion=data.competicion,
            fecha_jornada=data.fecha_jornada,
            estadio=data.estadio,
            fecha_hora=parse_fecha(data.fecha_hora),
            numero_partido=data.numero_partido,
            departamento=data.departamento,
            ciudad=data.ciudad,
            imagen_url=data.imagen_url,
        )
        db.add(partido)
        db.flush()

    # Crear asignaciones: buscar árbitro por nombre
    for asig in data.asignaciones:
        if not asig.get("nombre"):
            continue
        nombre_limpio = asig["nombre"].split("(")[0].strip().upper()
        arbitro = db.query(Arbitro).filter(Arbitro.nombre == nombre_limpio).first()
        if not arbitro:
            # Extraer categoría del paréntesis si existe
            cat_match = re.search(r"\(([^)]+)\)", asig["nombre"])
            categoria = cat_match.group(1).strip() if cat_match else "SIN CATEGORÍA"
            arbitro = Arbitro(nombre=nombre_limpio, categoria=categoria)
            db.add(arbitro)
            db.flush()

        a = AsignacionPartido(partido_id=partido.id, arbitro_id=arbitro.id, rol=asig["rol"])
        db.add(a)

    db.commit()
    conflictos = detectar_conflictos(partido.id, db)
    return {"id": partido.id, "conflictos": conflictos, "tiene_conflicto": len(conflictos) > 0}


@app.delete("/api/partidos/{partido_id}")
def eliminar_partido(partido_id: int, db: Session = Depends(get_db)):
    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido:
        raise HTTPException(status_code=404, detail="Partido no encontrado")
    db.delete(partido)
    db.commit()
    return {"ok": True}


# ─── Confirmaciones ──────────────────────────────────────────────────────────

@app.patch("/api/asignaciones/{asignacion_id}/confirmar")
def toggle_confirmacion(asignacion_id: int, db: Session = Depends(get_db)):
    a = db.query(AsignacionPartido).filter(AsignacionPartido.id == asignacion_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Asignación no encontrada")
    a.confirmado = not (a.confirmado or False)
    db.commit()
    return {"id": a.id, "confirmado": a.confirmado}


# ─── API Imagen OCR ───────────────────────────────────────────────────────────

@app.post("/api/procesar-imagen")
async def procesar_imagen(file: UploadFile = File(...)):
    contenido = await file.read()
    imagen_b64 = base64.b64encode(contenido).decode()
    media_type = file.content_type or "image/jpeg"

    # Guardar imagen en disco para usarla después en WhatsApp
    ext = file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "jpg"
    nombre_archivo = f"{int(datetime.utcnow().timestamp() * 1000)}.{ext}"
    ruta = os.path.join("static", "uploads", nombre_archivo)
    os.makedirs(os.path.join("static", "uploads"), exist_ok=True)
    with open(ruta, "wb") as f:
        f.write(contenido)
    imagen_url = f"/static/uploads/{nombre_archivo}"

    try:
        datos = extraer_datos_imagen(imagen_b64, media_type)
        return {"ok": True, "datos": datos, "imagen_url": imagen_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando imagen: {str(e)}")


@app.post("/api/procesar-imagen-simple")
async def subir_imagen_simple(file: UploadFile = File(...)):
    """Guarda la imagen sin procesarla con IA."""
    contenido = await file.read()
    ext = file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "jpg"
    nombre_archivo = f"{int(datetime.utcnow().timestamp() * 1000)}.{ext}"
    ruta = os.path.join("static", "uploads", nombre_archivo)
    os.makedirs(os.path.join("static", "uploads"), exist_ok=True)
    with open(ruta, "wb") as f:
        f.write(contenido)
    return {"imagen_url": f"/static/uploads/{nombre_archivo}"}


class ImagenUpdate(BaseModel):
    imagen_url: str


@app.patch("/api/partidos/{partido_id}/imagen")
def actualizar_imagen_partido(partido_id: int, data: ImagenUpdate, db: Session = Depends(get_db)):
    p = db.query(Partido).filter(Partido.id == partido_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Partido no encontrado")
    p.imagen_url = data.imagen_url
    db.commit()
    return {"ok": True}


# ─── API Conflictos y Sugerencias ────────────────────────────────────────────

@app.get("/api/conflictos")
def todos_los_conflictos(jornada_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(Partido)
    if jornada_id:
        q = q.filter(Partido.jornada_id == jornada_id)
    partidos = q.all()
    todos = []
    vistos = set()
    for p in partidos:
        for c in detectar_conflictos(p.id, db):
            clave = tuple(sorted([p.id, c["partido_conflicto_id"], c["arbitro_id"]]))
            if clave not in vistos:
                vistos.add(clave)
                c["partido_origen_id"] = p.id
                c["partido_origen_num"] = p.numero
                c["equipos_origen"] = f"{p.equipo_local} vs {p.equipo_visitante}"
                c["fecha_origen"] = p.fecha_hora.strftime("%d.%m.%Y %H:%M") if p.fecha_hora else ""
                c["estadio_origen"] = p.estadio or ""
                c["ciudad_origen"] = p.ciudad or ""
                c["competicion_origen"] = p.competicion or ""
                # Verificar si ya tiene reemplazo asignado
                reemplazo = db.query(Reemplazo).filter(
                    Reemplazo.arbitro_original_id == c["arbitro_id"],
                    Reemplazo.partido_id.in_([p.id, c["partido_conflicto_id"]])
                ).first()
                c["resuelto"] = reemplazo is not None
                c["reemplazo_nombre"] = reemplazo.arbitro_reemplazo.nombre if reemplazo else None
                c["reemplazo_partido_id"] = reemplazo.partido_id if reemplazo else None
                todos.append(c)
    return todos


@app.get("/api/sugerencias/{arbitro_id}/{partido_id}")
def sugerencias(arbitro_id: int, partido_id: int, db: Session = Depends(get_db)):
    return sugerir_reemplazos(arbitro_id, partido_id, db)


# ─── Reemplazos ──────────────────────────────────────────────────────────────

class ReemplazoCreate(BaseModel):
    partido_id: int
    arbitro_original_id: int
    arbitro_reemplazo_id: int
    rol: str


@app.get("/api/reemplazos")
def listar_reemplazos(jornada_id: Optional[int] = None, db: Session = Depends(get_db)):
    reemplazos = db.query(Reemplazo).order_by(Reemplazo.creado_en.desc()).all()
    if jornada_id:
        reemplazos = [r for r in reemplazos if r.partido.jornada_id == jornada_id]
    return [{
        "id": r.id,
        "partido_id": r.partido_id,
        "partido": f"{r.partido.equipo_local} vs {r.partido.equipo_visitante}",
        "fecha": r.partido.fecha_hora.strftime("%d.%m.%Y %H:%M") if r.partido.fecha_hora else "",
        "estadio": r.partido.estadio or "",
        "ciudad": r.partido.ciudad or "",
        "competicion": r.partido.competicion or "",
        "arbitro_original": r.arbitro_original.nombre,
        "arbitro_reemplazo": r.arbitro_reemplazo.nombre,
        "rol": r.rol,
    } for r in reemplazos]


@app.post("/api/reemplazos")
def crear_reemplazo(data: ReemplazoCreate, db: Session = Depends(get_db)):
    # Evitar duplicados
    existe = db.query(Reemplazo).filter(
        Reemplazo.partido_id == data.partido_id,
        Reemplazo.arbitro_original_id == data.arbitro_original_id,
    ).first()
    if existe:
        existe.arbitro_reemplazo_id = data.arbitro_reemplazo_id
        existe.rol = data.rol
        db.commit()
        return {"id": existe.id}
    r = Reemplazo(**data.dict())
    db.add(r)
    db.commit()
    db.refresh(r)
    return {"id": r.id}


@app.delete("/api/reemplazos/{reemplazo_id}")
def eliminar_reemplazo(reemplazo_id: int, db: Session = Depends(get_db)):
    r = db.query(Reemplazo).filter(Reemplazo.id == reemplazo_id).first()
    if r:
        db.delete(r)
        db.commit()
    return {"ok": True}


# ─── Exportar Excel ──────────────────────────────────────────────────────────

@app.get("/api/exportar/excel")
def exportar_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    wb = openpyxl.Workbook()

    # ── Hoja 1: Conflictos con reemplazos ──
    ws1 = wb.active
    ws1.title = "Conflictos y Reemplazos"

    header_fill = PatternFill("solid", fgColor="1a1d27")
    red_fill = PatternFill("solid", fgColor="e84040")
    orange_fill = PatternFill("solid", fgColor="f5a623")
    green_fill = PatternFill("solid", fgColor="2ecc71")
    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["ÁRBITRO", "TIPO", "PARTIDO 1", "FECHA 1", "CANCHA 1", "COMPETICIÓN 1",
               "PARTIDO 2", "FECHA 2", "CANCHA 2", "COMPETICIÓN 2", "REEMPLAZO ASIGNADO", "REEMPLAZA EN"]
    ws1.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3d7df7")
        cell.alignment = center

    # Obtener conflictos
    partidos = db.query(Partido).all()
    vistos = set()
    filas = []
    for p in partidos:
        for c in detectar_conflictos(p.id, db):
            clave = tuple(sorted([p.id, c["partido_conflicto_id"], c["arbitro_id"]]))
            if clave not in vistos:
                vistos.add(clave)
                # Buscar reemplazo asignado
                reemplazo = db.query(Reemplazo).filter(
                    Reemplazo.partido_id.in_([p.id, c["partido_conflicto_id"]]),
                    Reemplazo.arbitro_original_id == c["arbitro_id"]
                ).first()
                filas.append({
                    "arbitro": c["arbitro_nombre"],
                    "tipo": "SOLAPAMIENTO" if c["tipo"] == "solapamiento" else "MISMO DÍA",
                    "partido1": f"{p.equipo_local} vs {p.equipo_visitante}",
                    "fecha1": p.fecha_hora.strftime("%d.%m.%Y %H:%M") if p.fecha_hora else "",
                    "cancha1": f"{p.estadio or ''} {p.ciudad or ''}".strip(),
                    "comp1": p.competicion or "",
                    "partido2": c["equipos_conflicto"],
                    "fecha2": c["fecha_conflicto"],
                    "cancha2": f"{c.get('estadio_conflicto','') } {c.get('ciudad_conflicto','')}".strip(),
                    "comp2": c.get("competicion_conflicto", ""),
                    "reemplazo": reemplazo.arbitro_reemplazo.nombre if reemplazo else "Sin asignar",
                    "reemplaza_en": f"{reemplazo.partido.equipo_local} vs {reemplazo.partido.equipo_visitante}" if reemplazo else "",
                })

    for i, f in enumerate(filas, 2):
        row = [f["arbitro"], f["tipo"], f["partido1"], f["fecha1"], f["cancha1"], f["comp1"],
               f["partido2"], f["fecha2"], f["cancha2"], f["comp2"], f["reemplazo"], f["reemplaza_en"]]
        ws1.append(row)
        fill = PatternFill("solid", fgColor="2a1a1a") if f["tipo"] == "SOLAPAMIENTO" else PatternFill("solid", fgColor="2a2200")
        for col in range(1, len(row) + 1):
            cell = ws1.cell(row=i, column=col)
            cell.fill = fill
            if col == 11 and f["reemplazo"] != "Sin asignar":
                cell.font = Font(bold=True, color="2ecc71")
            elif col == 11:
                cell.font = Font(color="e84040")

    # Ajustar anchos
    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Hoja 2: Reemplazos confirmados ──
    ws2 = wb.create_sheet("Reemplazos Confirmados")
    h2 = ["ÁRBITRO REEMPLAZADO", "REEMPLAZADO POR", "ROL", "PARTIDO", "FECHA", "CANCHA", "CIUDAD", "COMPETICIÓN"]
    ws2.append(h2)
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2ecc71")
        cell.alignment = center

    reemplazos = db.query(Reemplazo).all()
    for r in reemplazos:
        ws2.append([
            r.arbitro_original.nombre,
            r.arbitro_reemplazo.nombre,
            r.rol,
            f"{r.partido.equipo_local} vs {r.partido.equipo_visitante}",
            r.partido.fecha_hora.strftime("%d.%m.%Y %H:%M") if r.partido.fecha_hora else "",
            r.partido.estadio or "",
            r.partido.ciudad or "",
            r.partido.competicion or "",
        ])
    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=conflictos_arbitros.xlsx"}
    )


# ─── Importar Excel ──────────────────────────────────────────────────────────

@app.post("/api/importar-excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Lee un Excel con columnas: NOMBRE, PARTIDO (o similar) y las variantes de rol.
    Formato esperado: columna ÁRBITRO con el nombre, columna ROL, columna FECHA, EQUIPO_LOCAL, EQUIPO_VISITANTE.
    Retorna preview de las columnas para que el usuario confirme el mapeo.
    """
    import openpyxl, io
    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v for v in row):
            filas.append([str(v).strip() if v is not None else "" for v in row])

    return {
        "columnas": headers,
        "total_filas": len(filas),
        "preview": filas[:3],  # Muestra las primeras 3 filas como ejemplo
    }


@app.post("/api/importar-excel/confirmar")
async def confirmar_importacion(
    file: UploadFile = File(...),
    col_nombre: int = 0,       # índice columna nombre árbitro
    col_rol: int = 1,          # índice columna rol
    col_fecha: int = 2,        # índice columna fecha
    col_local: int = 3,        # índice columna equipo local
    col_visitante: int = 4,    # índice columna equipo visitante
    col_estadio: int = -1,     # -1 = no existe
    col_ciudad: int = -1,
    db: Session = Depends(get_db)
):
    import openpyxl, io
    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    arbitros_creados = 0
    partidos_creados = 0
    errores = []

    def get_col(row, idx):
        try:
            return str(row[idx]).strip() if idx >= 0 and idx < len(row) else ""
        except:
            return ""

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = [str(v).strip() if v is not None else "" for v in row]
        nombre = get_col(row, col_nombre).upper()
        rol = get_col(row, col_rol)
        fecha_str = get_col(row, col_fecha)
        local = get_col(row, col_local).upper()
        visitante = get_col(row, col_visitante).upper()
        estadio = get_col(row, col_estadio)
        ciudad = get_col(row, col_ciudad)

        if not nombre or not local:
            continue

        # Crear o buscar árbitro
        arbitro = db.query(Arbitro).filter(Arbitro.nombre == nombre).first()
        if not arbitro:
            arbitro = Arbitro(nombre=nombre, categoria="")
            db.add(arbitro)
            db.flush()
            arbitros_creados += 1

        # Crear o buscar partido
        numero = f"XLS-{local[:3]}-{visitante[:3]}-{fecha_str}".replace(" ", "")
        partido = db.query(Partido).filter(Partido.numero == numero).first()
        if not partido:
            partido = Partido(
                numero=numero,
                equipo_local=local,
                equipo_visitante=visitante,
                competicion="",
                estadio=estadio,
                ciudad=ciudad,
                fecha_hora=parse_fecha(fecha_str),
            )
            db.add(partido)
            db.flush()
            partidos_creados += 1

        # Crear asignación si no existe
        existe = db.query(AsignacionPartido).filter(
            AsignacionPartido.partido_id == partido.id,
            AsignacionPartido.arbitro_id == arbitro.id,
        ).first()
        if not existe:
            db.add(AsignacionPartido(partido_id=partido.id, arbitro_id=arbitro.id, rol=rol or "Árbitro"))

    db.commit()
    return {"arbitros_creados": arbitros_creados, "partidos_creados": partidos_creados}


# ─── Estadísticas desde Excel ──────────────────────────────────────────────────

STATS_CACHE_PATH = "static/stats_cache.json"


def _norm(texto: str) -> str:
    """Quita tildes, mayúsculas, espacios extra. Normaliza para comparar."""
    s = unicodedata.normalize("NFD", texto)
    s = s.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s.upper().strip())


def _palabras(texto: str) -> list:
    return _norm(texto).split()


def _levenshtein(a: str, b: str) -> int:
    """Distancia de edición entre dos strings."""
    if len(a) < len(b):
        a, b = b, a
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(prev[j] + 1, curr[j-1] + 1, prev[j-1] + (ca != cb)))
        prev = curr
    return prev[-1]


def _fuzzy_word(w1: str, w2: str) -> bool:
    """True si las palabras son similares (typos de 1-2 chars en palabras largas)."""
    if w1 == w2:
        return True
    max_len = max(len(w1), len(w2))
    dist = _levenshtein(w1, w2)
    if max_len >= 7 and dist <= 2:
        return True
    if max_len >= 5 and dist <= 1:
        return True
    return False


def _score_match(excel_norm: str, db_norm: str) -> float:
    """
    Retorna score 0-1 de qué tan probable es que sean la misma persona.
    Estrategia: los 2 apellidos deben coincidir (con tolerancia typo) y al menos el primer nombre.
    """
    ep = _palabras(excel_norm)
    dp = _palabras(db_norm)
    if len(ep) < 2 or len(dp) < 2:
        return 0.0
    # Apellidos: exactos primero, luego fuzzy (typos leves)
    ap1_ok = ep[0] == dp[0] or _fuzzy_word(ep[0], dp[0])
    ap2_ok = ep[1] == dp[1] or _fuzzy_word(ep[1], dp[1])
    if not (ap1_ok and ap2_ok):
        return 0.0
    exact_ap = (ep[0] == dp[0] and ep[1] == dp[1])
    # Primer nombre (3ra palabra)
    if len(ep) >= 3 and len(dp) >= 3:
        n1, n2 = ep[2], dp[2]
        if n1 == n2:
            return 1.0 if exact_ap else 0.85
        if n1.startswith(n2) or n2.startswith(n1) or n1[0] == n2[0]:
            return 0.85 if exact_ap else 0.75
        # Fuzzy nombre también
        if _fuzzy_word(n1, n2):
            return 0.80 if exact_ap else 0.70
        return 0.3
    if len(ep) == 2:
        return 0.75 if exact_ap else 0.65
    return 0.0


def _vincular_con_db(stats_raw: dict, db: Session) -> dict:
    """
    Toma el dict {nombre_excel: {...}} del Excel y fusiona entradas que
    corresponden al mismo árbitro de la DB.
    Retorna lista con campo arbitro_id, nombre_db, nombre_excel, linked.
    """
    arbitros_db = db.query(Arbitro).all()

    # Precalcular versión normalizada de cada árbitro DB (sin coma)
    db_info = []
    for a in arbitros_db:
        nombre_sin_coma = a.nombre.replace(", ", " ").replace(",", " ")
        db_info.append({
            "id": a.id,
            "nombre_db": a.nombre,
            "norm": _norm(nombre_sin_coma),
        })

    # Para cada entrada del Excel, buscar el mejor match en DB
    mapping = {}  # nombre_excel → arbitro_id (o None)
    for nombre_excel in stats_raw:
        norm_excel = _norm(nombre_excel)
        best_id, best_score = None, 0.0
        for d in db_info:
            s = _score_match(norm_excel, d["norm"])
            if s > best_score:
                best_score, best_id = s, d["id"]
        mapping[nombre_excel] = best_id if best_score >= 0.75 else None

    # Agrupar por arbitro_id (merge) o por nombre_excel si no vinculado
    merged: dict = {}  # clave → acumulador

    for nombre_excel, s in stats_raw.items():
        arb_id = mapping[nombre_excel]
        if arb_id is not None:
            clave = f"db:{arb_id}"
            arb = next(d for d in db_info if d["id"] == arb_id)
            if clave not in merged:
                merged[clave] = {
                    "nombre": arb["nombre_db"],
                    "arbitro_id": arb_id,
                    "linked": True,
                    "nombres_excel": [],
                    "total": 0, "arbitro_principal": 0,
                    "primer_asistente": 0, "segundo_asistente": 0,
                    "cuarto_arbitro": 0, "torneos": defaultdict(int),
                    "jornadas": set(),
                }
            merged[clave]["nombres_excel"].append(nombre_excel)
            merged[clave]["total"]             += s["total"]
            merged[clave]["arbitro_principal"] += s["arbitro_principal"]
            merged[clave]["primer_asistente"]  += s["primer_asistente"]
            merged[clave]["segundo_asistente"] += s["segundo_asistente"]
            merged[clave]["cuarto_arbitro"]    += s["cuarto_arbitro"]
            for t, n in s["torneos"].items():
                merged[clave]["torneos"][t] += n
            merged[clave]["jornadas"].update(s["jornadas"])
        else:
            clave = f"excel:{nombre_excel}"
            merged[clave] = {
                "nombre": nombre_excel,
                "arbitro_id": None,
                "linked": False,
                "nombres_excel": [nombre_excel],
                **s,
                "torneos": dict(s["torneos"]),
                "jornadas": s["jornadas"],
            }

    result = []
    for entry in merged.values():
        result.append({
            "nombre":            entry["nombre"],
            "arbitro_id":        entry["arbitro_id"],
            "linked":            entry["linked"],
            "nombres_excel":     entry["nombres_excel"],
            "total":             entry["total"],
            "arbitro_principal": entry["arbitro_principal"],
            "primer_asistente":  entry["primer_asistente"],
            "segundo_asistente": entry["segundo_asistente"],
            "cuarto_arbitro":    entry["cuarto_arbitro"],
            "torneos":           dict(entry["torneos"]),
            "jornadas":          len(entry["jornadas"]),
        })

    result.sort(key=lambda x: (-x["total"], x["nombre"]))
    return result


@app.post("/api/stats/importar")
async def importar_stats_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    filename = file.filename or "Excel"

    stats_raw = defaultdict(lambda: {
        "total": 0, "arbitro_principal": 0,
        "primer_asistente": 0, "segundo_asistente": 0,
        "cuarto_arbitro": 0, "torneos": defaultdict(int), "jornadas": set(),
    })

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        torneo = str(row[1] or "").strip()
        fecha  = row[3]
        roles_nombres = [
            ("arbitro_principal",  str(row[8]  or "").strip()),
            ("primer_asistente",   str(row[9]  or "").strip()),
            ("segundo_asistente",  str(row[10] or "").strip()),
            ("cuarto_arbitro",     str(row[11] or "").strip()),
        ]
        for rol_key, nombre in roles_nombres:
            if nombre:
                stats_raw[nombre]["total"] += 1
                stats_raw[nombre][rol_key] += 1
                stats_raw[nombre]["torneos"][torneo] += 1
                stats_raw[nombre]["jornadas"].add(fecha)

    data = _vincular_con_db(stats_raw, db)
    cache = {"filename": filename, "data": data}
    with open(STATS_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    return cache


@app.get("/api/stats")
def obtener_stats():
    if not os.path.exists(STATS_CACHE_PATH):
        return {"filename": None, "data": []}
    with open(STATS_CACHE_PATH, encoding="utf-8") as f:
        return json.load(f)


@app.post("/api/stats/revincular")
def revincular_stats(db: Session = Depends(get_db)):
    """Re-procesa el cache actual contra los árbitros actuales en la DB."""
    if not os.path.exists(STATS_CACHE_PATH):
        raise HTTPException(status_code=404, detail="No hay stats cargadas")
    with open(STATS_CACHE_PATH, encoding="utf-8") as f:
        cache = json.load(f)

    # Reconstruir stats_raw desde el cache actual (usando nombres_excel)
    stats_raw = defaultdict(lambda: {
        "total": 0, "arbitro_principal": 0,
        "primer_asistente": 0, "segundo_asistente": 0,
        "cuarto_arbitro": 0, "torneos": defaultdict(int), "jornadas": set(),
    })
    for entry in cache["data"]:
        for nex in entry.get("nombres_excel", [entry["nombre"]]):
            stats_raw[nex]["total"]             += entry["total"]
            stats_raw[nex]["arbitro_principal"] += entry["arbitro_principal"]
            stats_raw[nex]["primer_asistente"]  += entry["primer_asistente"]
            stats_raw[nex]["segundo_asistente"] += entry["segundo_asistente"]
            stats_raw[nex]["cuarto_arbitro"]    += entry["cuarto_arbitro"]
            for t, n in entry["torneos"].items():
                stats_raw[nex]["torneos"][t] += n

    data = _vincular_con_db(stats_raw, db)
    cache["data"] = data
    with open(STATS_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    return cache


# ─── Restaurar DB (uso único para migración) ──────────────────────────────────
from database import _db_path as _DB_FILE_PATH
from fastapi.responses import HTMLResponse

@app.get("/admin/restaurar-db", response_class=HTMLResponse)
async def restaurar_db_form():
    return """
    <html><body style="font-family:sans-serif;max-width:400px;margin:60px auto;padding:20px">
    <h2>Restaurar base de datos</h2>
    <form method="post" enctype="multipart/form-data">
      <p><label>Clave:<br><input type="password" name="clave" style="width:100%;padding:8px;margin-top:4px"></label></p>
      <p><label>Archivo .db:<br><input type="file" name="file" accept=".db" style="margin-top:4px"></label></p>
      <button type="submit" style="background:#0099CC;color:#fff;border:none;padding:10px 20px;border-radius:6px;cursor:pointer">Subir DB</button>
    </form>
    </body></html>
    """

@app.post("/admin/restaurar-db")
async def restaurar_db(file: UploadFile = File(...), clave: str = ""):
    if clave.strip() != os.getenv("ADMIN_CLAVE", "").strip():
        raise HTTPException(status_code=403, detail="Clave incorrecta")
    contenido = await file.read()
    with open(_DB_FILE_PATH, "wb") as f:
        f.write(contenido)
    return HTMLResponse("<html><body style='font-family:sans-serif;max-width:400px;margin:60px auto'><h2>✅ Base de datos restaurada correctamente</h2><a href='/'>Ir a la app</a></body></html>")
